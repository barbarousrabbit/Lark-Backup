"""
Data comparison module
Used to compare backup data from different dates and detect data changes
"""

import os
import glob
import json
import time
import logging
from collections import Counter
from datetime import datetime, timedelta
import openpyxl
from typing import Dict, List, Optional, Tuple

from config import config

class DataComparator:
    """Data comparator for comparing backup files from different dates"""

    def __init__(self, backup_dir: str):
        """
        Initialize the data comparator

        Args:
            backup_dir: Directory where backup files are stored
        """
        self.backup_dir = backup_dir
        self.comparison_cache_file = os.path.join(
            config.get_program_dir(), "data_comparison_cache.json"
        )

    def get_backup_file_path(self, date_str: str) -> str:
        """
        Resolve a date to its backup file.

        Prefers the canonical name, but file_manager falls back to a
        timestamped name ("... {date}_HHMMSS.xlsx") or the program-dir
        "backup" folder when the canonical path is locked/unwritable. Return
        whichever real file exists (newest timestamped match) so comparison
        and data-loss detection keep working after a fallback save. When none
        exists, return the canonical path so file_exists() reports missing.

        Args:
            date_str: Date string in YYYY-MM-DD format

        Returns:
            File path
        """
        filename = config.BACKUP_FILENAME_TEMPLATE.format(date=date_str)
        canonical = os.path.join(self.backup_dir, filename)
        if os.path.exists(canonical):
            return canonical

        base, ext = os.path.splitext(filename)
        pattern = f"{base}_*{ext}"
        matches = []
        for search_dir in (self.backup_dir, os.path.join(config.get_program_dir(), "backup")):
            try:
                matches.extend(glob.glob(os.path.join(search_dir, pattern)))
            except OSError:
                continue
        if matches:
            try:
                return max(matches, key=os.path.getmtime)
            except OSError:
                return matches[0]
        return canonical

    def file_exists(self, date_str: str) -> bool:
        """
        Check whether the backup file for a given date exists

        Args:
            date_str: Date string

        Returns:
            Whether the file exists
        """
        file_path = self.get_backup_file_path(date_str)
        return os.path.exists(file_path)

    @staticmethod
    def _row_key(row) -> Optional[str]:
        """
        Serialize one worksheet row into its comparison key.

        Returns None for rows without content so callers can skip them.
        Blank rows occur in the MIDDLE of real Lark exports — a blank row
        must never be treated as end-of-data.
        """
        if not any(cell.value is not None for cell in row):
            return None
        values = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
        while values and values[-1] == '':
            values.pop()
        return '|'.join(values) if values else None

    def get_all_sheet_counters(self, file_path: str) -> Dict[str, Counter]:
        """
        Build a data-row Counter for EVERY worksheet in one streaming pass.

        Two hard requirements learned from production files:
        - Lark's exporter writes a bogus `<dimension ref="A1"/>` on every
          sheet. openpyxl read-only mode trusts it, so iter_rows() yields
          ZERO rows unless reset_dimensions() is called first.
        - Sheets contain interior blank rows, so every row is scanned — an
          early break on the first blank row silently under-counts.

        Raises on an unreadable file. Callers must treat that as
        "comparison impossible" — degrading to an empty Counter would count
        every row of the other file as deleted and fire a false data-loss
        alert.

        Args:
            file_path: Path to the Excel file

        Returns:
            Mapping from worksheet name to its data-row Counter
        """
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            counters: Dict[str, Counter] = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet.reset_dimensions()
                counter: Counter = Counter()
                for row in sheet.iter_rows(min_row=2):
                    key = self._row_key(row)
                    if key is not None:
                        counter[key] += 1
                counters[sheet_name] = counter
            return counters
        finally:
            if workbook:
                workbook.close()

    def compare_two_dates(self, date1_str: str, date2_str: str) -> Tuple[Dict[str, int], List[str]]:
        """
        Compare backup data from two dates.
        Data-loss WARNINGS fire on net row-count decline (the sheet shrank);
        the differences dict still carries Counter-level deleted/added detail
        for the report.

        Args:
            date1_str: First date (earlier)
            date2_str: Second date (later)

        Returns:
            (differences dict, list of warning messages)
        """
        warnings = []
        differences = {}

        # Check whether files exist
        if not self.file_exists(date1_str):
            logging.warning(f"⚠️ File for {date1_str} does not exist")
            warnings.append(f"Backup file for {date1_str} does not exist")
            return {}, warnings

        if not self.file_exists(date2_str):
            logging.warning(f"⚠️ File for {date2_str} does not exist")
            warnings.append(f"Backup file for {date2_str} does not exist")
            return {}, warnings

        # Get paths for both files
        file1_path = self.get_backup_file_path(date1_str)
        file2_path = self.get_backup_file_path(date2_str)

        # One streaming pass per file builds every sheet's Counter
        # (measured ~3.5 min per 200+ MB export)
        start_time = time.time()
        try:
            counters1 = self.get_all_sheet_counters(file1_path)
            counters2 = self.get_all_sheet_counters(file2_path)
        except Exception as e:
            # An unreadable file aborts the comparison with a warning —
            # never degrade to "empty file", which would report every row
            # of the other file as deleted and fire a false data-loss alert.
            logging.error(f"❌ Comparison aborted — cannot read backup file: {e}")
            warnings.append(f"Comparison could not run: {e}")
            return {}, warnings

        # Compare data content for each worksheet
        all_sheets = set(counters1.keys()) | set(counters2.keys())

        for sheet_name in all_sheets:
            data1 = counters1.get(sheet_name, Counter())
            data2 = counters2.get(sheet_name, Counter())
            count1 = sum(data1.values())
            count2 = sum(data2.values())

            if count1 == 0 and count2 == 0:
                continue  # Neither file has data for this sheet

            # Calculate data changes (Counter subtraction discards zeros/negatives, keeping only surplus)
            deleted_counter = data1 - data2   # rows in data1 not fully covered by data2
            added_counter = data2 - data1     # rows in data2 not fully covered by data1

            deleted_count = sum(deleted_counter.values())
            added_count = sum(added_counter.values())
            net_change = count2 - count1

            # Record detailed change information
            if deleted_count > 0 or added_count > 0:
                differences[sheet_name] = {
                    'before': count1,
                    'after': count2,
                    'difference': net_change,
                    'deleted_count': deleted_count,
                    'added_count': added_count
                }

                logging.info(f"📊 Sheet '{sheet_name}': {count1} → {count2} (deleted: {deleted_count}, added: {added_count})")

                # Data-loss warning on NET decline only. Counter "deleted"
                # also counts rows that merely CHANGED (volatile computed
                # columns rewrite thousands of rows every day), so alerting
                # on it would cry wolf daily; a real mass deletion shows up
                # as the sheet shrinking.
                net_loss = count1 - count2
                if net_loss > config.ALERT_DELETED_ROW_THRESHOLD:
                    warning_msg = f"⚠️ Sheet '{sheet_name}' shrank by {net_loss} rows ({count1} → {count2})"
                    warnings.append(warning_msg)
                    logging.warning(warning_msg)

        logging.info(f"🔍 Compared {len(all_sheets)} sheets in {time.time() - start_time:.0f}s")
        return differences, warnings

    def compare_with_previous_date(self, current_date_str: str) -> Tuple[Dict[str, int], List[str]]:
        """
        Compare the current date's backup with the previous day

        Args:
            current_date_str: Current date string

        Returns:
            (differences dict, list of warning messages)
        """
        current_date = datetime.strptime(current_date_str, "%Y-%m-%d")
        previous_date = current_date - timedelta(days=1)
        previous_date_str = previous_date.strftime("%Y-%m-%d")

        logging.info(f"🔍 Comparing {previous_date_str} with {current_date_str}")
        return self.compare_two_dates(previous_date_str, current_date_str)

    def save_comparison_result(self, date_str: str, comparison_result: Dict):
        """
        Save comparison result to the cache file

        Args:
            date_str: Date string
            comparison_result: Comparison result
        """
        try:
            # Load existing cache
            if os.path.exists(self.comparison_cache_file):
                with open(self.comparison_cache_file, 'r', encoding='utf-8') as f:
                    cache = json.load(f)
            else:
                cache = {}

            # Update cache
            cache[date_str] = {
                'timestamp': datetime.now().isoformat(),
                'result': comparison_result
            }

            # Keep only the last 30 days of records
            cutoff_date = datetime.now() - timedelta(days=30)
            cache = {
                k: v for k, v in cache.items()
                if datetime.fromisoformat(v['timestamp']) > cutoff_date
            }

            # Save cache
            with open(self.comparison_cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)

        except Exception as e:
            logging.error(f"❌ Error saving comparison cache: {e}")

# Global singleton
data_comparator = DataComparator(config.DOWNLOAD_DIR)

def init_data_comparator(backup_dir: str) -> DataComparator:
    """Initialize and return the data comparator (returns the global singleton, backup_dir parameter is ignored)"""
    return data_comparator
